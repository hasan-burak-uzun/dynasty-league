#!/usr/bin/env python3
"""
update_rosters.py — Sync current Yahoo NBA Fantasy rosters into the salary Excel file.

Usage:
    uv run update_rosters.py
    uv run update_rosters.py --source "C:/path/to/salaries.xlsx"
    uv run update_rosters.py --output "C:/Users/.../Google Drive/dynasty_rosters.xlsx"
    uv run update_rosters.py --debug
"""

import json
import re
import sys
import argparse
import unicodedata
from pathlib import Path
from datetime import date

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import requests
from rapidfuzz import fuzz, process as rfprocess
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ── Config ─────────────────────────────────────────────────────────────────────

LEAGUE_ID       = "2197"
LEAGUE_KEY      = f"nba.l.{LEAGUE_ID}"
API_BASE        = "https://pub-api-rw.fantasysports.yahoo.com/fantasy/v2"
CRUMB_URL       = "https://graviton-user-gateway.media.yahoo.com/api/auth/v2/crumb"
UA              = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

DEFAULT_SALARY    = 7.5
FUZZY_THRESHOLD   = 82          # lower = more lenient name matching
COOKIES_FILE      = Path(__file__).parent / "cookies.txt"
DEBUG_ROSTER_FILE = Path(__file__).parent / "debug_roster.json"

SALARY_SHEET      = "Over 7,5"
SUFFIXES          = {"jr", "sr", "ii", "iii", "iv", "v"}
YELLOW_FILL       = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")

# Google Drive candidate paths (checked in order)
GDRIVE_CANDIDATES = [
    Path.home() / "Google Drive",
    Path.home() / "My Drive",
    Path("G:") / "My Drive",
    Path("G:/"),
]

# ── Cookies ────────────────────────────────────────────────────────────────────

def load_cookie_header() -> str:
    if not COOKIES_FILE.exists():
        print(f"ERROR: {COOKIES_FILE} not found.")
        print("Paste your Yahoo Cookie header (from F12 → Network tab) into that file.")
        sys.exit(1)
    return COOKIES_FILE.read_text(encoding="utf-8").strip()

# ── Yahoo API ──────────────────────────────────────────────────────────────────

def _get_crumb(cookie_header: str) -> str | None:
    try:
        r = requests.get(
            CRUMB_URL, params={"appId": "yahoo_fp"},
            headers={"Cookie": cookie_header, "User-Agent": UA}, timeout=10,
        )
        d = r.json()
        return d.get("crumb") or d.get("value") or d.get("crumb_b64")
    except Exception:
        return None


def _api_get(path: str, cookie_header: str, crumb: str | None, params: dict | None = None) -> dict:
    p = {"format": "json", **(params or {})}
    if crumb:
        p["crumb"] = crumb
    r = requests.get(
        f"{API_BASE}/{path}",
        headers={"Cookie": cookie_header, "User-Agent": UA},
        params=p, timeout=20,
    )
    r.raise_for_status()
    return r.json()


def fetch_actual_league_key(cookie_header: str, crumb: str | None) -> str:
    data = _api_get(f"league/{LEAGUE_KEY}/standings", cookie_header, crumb)
    return data["fantasy_content"]["league"][0]["league_key"]


def fetch_rosters(cookie_header: str, crumb: str | None, actual_key: str, debug: bool) -> dict:
    data = _api_get(f"league/{actual_key}/teams/roster", cookie_header, crumb)
    if debug:
        DEBUG_ROSTER_FILE.write_text(json.dumps(data, indent=2), encoding="utf-8")
        print(f"  Debug: raw roster response → {DEBUG_ROSTER_FILE}")
    return data

# ── Name Utilities ─────────────────────────────────────────────────────────────

def _strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


def normalize_for_match(name: str) -> str:
    """Lowercase, strip accents, remove apostrophes/periods, strip NBA team code suffix."""
    name = _strip_accents(str(name)).replace("\xa0", " ")
    name = re.sub(r"['.`]", "", name)
    name = re.sub(r"\s+[A-Z]{2,4}$", "", name.strip())   # strip "DEN", "OKC", etc.
    return name.lower().strip()


def abbreviate_name(full_name: str) -> str:
    """
    'Giannis Antetokounmpo' → 'G. Antetokounmpo'
    'Michael Porter Jr.'   → 'M. Porter Jr'
    'LaMelo Ball'          → 'L. Ball'
    """
    parts = full_name.strip().split()
    if not parts:
        return full_name

    # Peel off suffix (Jr., III, etc.)
    suffix = ""
    if len(parts) > 1 and parts[-1].rstrip(".").lower() in SUFFIXES:
        suffix = " " + parts[-1].rstrip(".")
        parts = parts[:-1]

    if len(parts) == 1:
        return parts[0] + suffix

    first_init = parts[0][0].upper() + "."
    last = " ".join(parts[1:])
    return f"{first_init} {last}{suffix}"


def format_player_cell(full_name: str, nba_team: str, positions: list[str]) -> str:
    """Return 'F. Surname TEAM – POS1,POS2' (en-dash separator)."""
    abbrev   = abbreviate_name(full_name)
    team_str = (nba_team or "???").upper()
    pos_str  = ",".join(p for p in positions if p not in {"BN", "IL", "IL+", "NA"})
    sep      = "\u2013"   # en-dash, matching the original file
    return f"{abbrev} {team_str} {sep} {pos_str}"

# ── Salary Map ─────────────────────────────────────────────────────────────────

def load_salary_map(wb) -> dict[str, float]:
    """Build {normalized_name: salary} from the 'Over 7,5' sheet."""
    sheet = wb[SALARY_SHEET]
    salary_map: dict[str, float] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_name, raw_sal = (row + (None, None))[:2]
        if not raw_name or raw_sal is None:
            continue
        try:
            sal = float(raw_sal)
        except (ValueError, TypeError):
            continue
        norm = normalize_for_match(str(raw_name))
        if norm:
            salary_map[norm] = sal
    return salary_map


def match_salary(yahoo_name: str, salary_map: dict[str, float], debug: bool = False) -> tuple[float, str]:
    """Return (salary, match_status) for a Yahoo player name."""
    norm = normalize_for_match(yahoo_name)

    # Exact
    if norm in salary_map:
        return salary_map[norm], "exact"

    # Fuzzy
    result = rfprocess.extractOne(norm, salary_map.keys(), scorer=fuzz.token_sort_ratio)
    if result and result[1] >= FUZZY_THRESHOLD:
        if debug:
            print(f"    fuzzy: '{yahoo_name}' → '{result[0]}' ({result[1]}%)")
        return salary_map[result[0]], f"fuzzy({result[1]}%)"

    return DEFAULT_SALARY, "not_found"

# ── Parse Yahoo Roster Response ────────────────────────────────────────────────

def _extract_eligible_positions(player_meta: list) -> list[str]:
    """Pull eligible positions from player metadata list."""
    for item in player_meta:
        if not isinstance(item, dict):
            continue
        ep = item.get("eligible_positions")
        if ep is None:
            continue
        # Could be {"count":2, "0":{"position":"PG"}, "1":{"position":"SG"}}
        # or a list [{"position":"PG"}, {"position":"SG"}]
        if isinstance(ep, dict):
            positions = []
            count = int(ep.get("count", len(ep)))
            for i in range(count):
                p = ep.get(str(i), {})
                if isinstance(p, dict) and "position" in p:
                    positions.append(p["position"])
            return positions
        if isinstance(ep, list):
            return [p["position"] if isinstance(p, dict) else str(p) for p in ep]
    return []


def _iter_numeric_keys(d: dict):
    """Yield values for keys "0", "1", "2", ... as long as they exist."""
    i = 0
    while str(i) in d:
        yield d[str(i)]
        i += 1


def parse_roster_response(data: dict) -> list[dict]:
    """
    Returns list of {team_name: str, players: [{name, nba_team, positions}]}.

    Actual API structure:
      league[1]["teams"]["0"]["team"][0]         → team metadata list
      league[1]["teams"]["0"]["team"][1]["roster"]["0"]["players"]["0"]["player"]
    """
    try:
        teams_block = data["fantasy_content"]["league"][1]["teams"]
    except (KeyError, IndexError, TypeError) as e:
        print(f"ERROR: unexpected API structure — {e}")
        print("Re-run with --debug to inspect debug_roster.json")
        sys.exit(1)

    result = []

    for team_entry in _iter_numeric_keys(teams_block):
        team_arr = team_entry.get("team", [])
        if not team_arr:
            continue

        # Team name from metadata list (list of single-key dicts)
        meta      = team_arr[0] if team_arr else []
        team_name = next(
            (d["name"] for d in meta if isinstance(d, dict) and "name" in d),
            f"Team ?",
        )

        # Players: roster → "0" → players → "0".."N"
        players_block = {}
        if len(team_arr) > 1:
            roster = team_arr[1].get("roster", {})
            players_block = roster.get("0", {}).get("players", {})

        players = []
        for player_entry in _iter_numeric_keys(players_block):
            pe          = player_entry.get("player", [])
            player_meta = pe[0] if pe else []

            # Full name is nested: {"name": {"full": "Reed Sheppard", ...}}
            name_obj  = next(
                (d["name"] for d in player_meta if isinstance(d, dict) and isinstance(d.get("name"), dict)),
                None,
            )
            full_name = name_obj["full"] if name_obj else None

            nba_team  = next(
                (d["editorial_team_abbr"] for d in player_meta
                 if isinstance(d, dict) and "editorial_team_abbr" in d),
                "???",
            )

            # display_position ("PG,SG") is cleaner than eligible_positions
            display_pos = next(
                (d["display_position"] for d in player_meta
                 if isinstance(d, dict) and "display_position" in d),
                None,
            )
            positions = display_pos.split(",") if display_pos else _extract_eligible_positions(player_meta)

            if full_name:
                players.append({"name": full_name, "nba_team": nba_team, "positions": positions})

        result.append({"name": team_name, "players": players})

    return result

# ── Excel Update ───────────────────────────────────────────────────────────────

def _get_team_headers(wb) -> dict[str, str]:
    """Return {sheet_name: A1_value} for every non-salary sheet."""
    headers = {}
    for name in wb.sheetnames:
        if name == SALARY_SHEET:
            continue
        a1 = wb[name].cell(1, 1).value
        if a1:
            headers[name] = str(a1).strip()
    return headers


def _match_sheet(yahoo_name: str, team_headers: dict[str, str]) -> str | None:
    """Match Yahoo team name to Excel sheet.

    Strategy 1 — full-name fuzzy match.
    Strategy 2 — first-word (city) exact match as fallback, to handle cases
                 like 'Artvin 99ers' vs 'Artvin Ninetyniners'.
    """
    result = rfprocess.extractOne(yahoo_name, team_headers.values(), scorer=fuzz.token_sort_ratio)
    if result and result[1] >= 55:
        for sheet_name, header in team_headers.items():
            if header == result[0]:
                return sheet_name

    # Fallback: match on first word only (city/prefix)
    yahoo_city = yahoo_name.split()[0].lower() if yahoo_name else ""
    for sheet_name, header in team_headers.items():
        if yahoo_city and header.split()[0].lower() == yahoo_city:
            return sheet_name

    return None


def _find_player_rows(sheet) -> tuple[int, int]:
    """Return (first_row, last_row) of player data (stops at TOTAL row)."""
    first = 3
    last  = first
    for row_num in range(3, 25):
        val = str(sheet.cell(row_num, 1).value or "").strip().upper()
        if val == "TOTAL":
            break
        last = row_num
    return first, last


def update_excel(
    source_path: Path,
    yahoo_teams: list[dict],
    salary_map: dict[str, float],
    output_path: Path,
    debug: bool,
) -> None:
    wb           = load_workbook(source_path)
    team_headers = _get_team_headers(wb)

    matched_count   = 0
    unmatched_names = []

    for yt in yahoo_teams:
        sheet_name = _match_sheet(yt["name"], team_headers)
        if not sheet_name:
            print(f"  WARNING: no Excel sheet matched Yahoo team '{yt['name']}'")
            continue

        sheet = wb[sheet_name]
        first_row, last_row = _find_player_rows(sheet)
        players = yt["players"]
        matched_count += 1

        if debug:
            print(f"  '{yt['name']}' → sheet '{sheet_name}' "
                  f"(rows {first_row}–{last_row}, {len(players)} players)")

        for row_idx in range(first_row, last_row + 1):
            p_idx = row_idx - first_row
            cell_a = sheet.cell(row_idx, 1)
            cell_b = sheet.cell(row_idx, 2)

            if p_idx < len(players):
                p       = players[p_idx]
                salary, status = match_salary(p["name"], salary_map, debug)
                if status == "not_found":
                    unmatched_names.append(p["name"])

                cell_a.value = format_player_cell(p["name"], p["nba_team"], p["positions"])
                cell_b.value = salary
                cell_a.fill  = YELLOW_FILL
            else:
                # Roster shorter than expected — clear the cell
                cell_a.value = ""
                cell_b.value = DEFAULT_SALARY
                cell_a.fill  = YELLOW_FILL

    print(f"  Matched {matched_count}/{len(yahoo_teams)} Yahoo teams to Excel sheets")

    if unmatched_names:
        unique = sorted(set(unmatched_names))
        print(f"  {len(unique)} player(s) not found in salary list -> assigned ${DEFAULT_SALARY}M:")
        for n in unique:
            print(f"    - {n}")

    # Sort team tabs by total salary (highest → lowest); keep salary sheet first
    def _sheet_total(sheet_name: str) -> float:
        s = wb[sheet_name]
        total = 0.0
        for row_num in range(3, 25):
            val = str(s.cell(row_num, 1).value or "").strip().upper()
            if val == "TOTAL":
                break
            sal = s.cell(row_num, 2).value
            try:
                total += float(sal or 0)
            except (ValueError, TypeError):
                pass
        return total

    team_sheet_names = [n for n in wb.sheetnames if n != SALARY_SHEET]
    team_sheet_names.sort(key=_sheet_total, reverse=True)

    # Rebuild sheet order: salary sheet first, then teams by total salary
    ordered = [wb[SALARY_SHEET]] + [wb[n] for n in team_sheet_names]
    wb._sheets = ordered

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"  Saved → {output_path}")

# ── Main ────────────────────────────────────────────────────────────────────────

def find_google_drive() -> Path | None:
    for p in GDRIVE_CANDIDATES:
        if p.exists() and p.is_dir():
            return p
    return None


def main() -> None:
    parser = argparse.ArgumentParser(description="Update dynasty league roster Excel")
    parser.add_argument("--source",  help="Path to the salary Excel file")
    parser.add_argument("--output",  help="Output file path (default: Google Drive or local)")
    parser.add_argument("--debug",   action="store_true", help="Verbose output + save raw API response")
    args = parser.parse_args()

    # Source file
    source_path = (
        Path(args.source) if args.source
        else Path(r"C:\Users\Hasan Burak Uzun\Downloads\salaries_251220 (1).xlsx")
    )
    if not source_path.exists():
        print(f"ERROR: source file not found: {source_path}")
        print("Use --source to specify its path.")
        sys.exit(1)

    # Output file
    today_str = date.today().strftime("%Y-%m-%d")
    fname     = f"dynasty_rosters_{today_str}.xlsx"
    if args.output:
        output_path = Path(args.output)
    else:
        gdrive = find_google_drive()
        if gdrive:
            output_path = gdrive / fname
            print(f"Google Drive: {gdrive}")
        else:
            output_path = Path(__file__).parent / fname
            print("Google Drive not found — saving locally")

    print("=== Dynasty League Roster Update ===")
    print("Loading cookies ...")
    cookie_header = load_cookie_header()

    print("Connecting to Yahoo ...")
    crumb = _get_crumb(cookie_header)
    actual_key = fetch_actual_league_key(cookie_header, crumb)
    print(f"  League key: {actual_key}")

    print("Fetching rosters ...")
    raw_data    = fetch_rosters(cookie_header, crumb, actual_key, args.debug)
    yahoo_teams = parse_roster_response(raw_data)
    total_players = sum(len(t["players"]) for t in yahoo_teams)
    print(f"  Got {len(yahoo_teams)} teams, {total_players} players")

    print("Loading salary data ...")
    wb_src     = load_workbook(source_path)
    salary_map = load_salary_map(wb_src)
    print(f"  {len(salary_map)} players in salary sheet")

    print("Building updated roster file ...")
    update_excel(source_path, yahoo_teams, salary_map, output_path, args.debug)

    print(f"\nDone.  Open: {output_path}")


if __name__ == "__main__":
    main()
